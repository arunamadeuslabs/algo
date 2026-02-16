"""
Visualization module for backtest results.
Generates charts for equity curve, trades, signals, and performance metrics.
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import FancyBboxPatch
import os

from config import RESULTS_DIR


def ensure_results_dir():
    os.makedirs(RESULTS_DIR, exist_ok=True)


def plot_signals_chart(data: pd.DataFrame, trades, save_path: str = None):
    """
    Plot price chart with EMAs, crossover signals (+ signs), and trade entries/exits.
    """
    ensure_results_dir()
    if save_path is None:
        save_path = os.path.join(RESULTS_DIR, "signals_chart.png")

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(20, 12), height_ratios=[3, 1],
                                     sharex=True, facecolor='#1a1a2e')

    for ax in [ax1, ax2]:
        ax.set_facecolor('#16213e')
        ax.tick_params(colors='white')
        ax.spines['bottom'].set_color('#444')
        ax.spines['top'].set_color('#444')
        ax.spines['left'].set_color('#444')
        ax.spines['right'].set_color('#444')

    # --- Price & EMAs ---
    ax1.plot(data.index, data['close'], color='#e0e0e0', linewidth=0.8, alpha=0.7, label='Close')
    ax1.plot(data.index, data['ema_fast'], color='#ff9800', linewidth=1.5, label=f'EMA Fast (Orange)')
    ax1.plot(data.index, data['ema_slow'], color='#4caf50', linewidth=1.5, label=f'EMA Slow (Green)')

    # --- Plus Signs for crossovers ---
    bull_crosses = data[data['crossover_bull'] == True]
    bear_crosses = data[data['crossover_bear'] == True]

    ax1.scatter(bull_crosses.index, bull_crosses['low'] - 20,
                marker='+', color='#00e676', s=200, linewidths=3,
                label='+ Bullish Cross', zorder=5)
    ax1.scatter(bear_crosses.index, bear_crosses['high'] + 20,
                marker='+', color='#ff1744', s=200, linewidths=3,
                label='+ Bearish Cross', zorder=5)

    # --- Trade Entries & Exits ---
    for trade in trades:
        color = '#00e676' if trade.direction.value == 1 else '#ff1744'
        marker_entry = '^' if trade.direction.value == 1 else 'v'

        ax1.scatter(trade.entry_time, trade.entry_spot,
                    marker=marker_entry, color=color, s=150, zorder=6, edgecolors='white')

        if trade.exit_time:
            ax1.scatter(trade.exit_time, trade.exit_spot,
                        marker='x', color='#ffeb3b', s=100, zorder=6, linewidths=2)
            # Draw trade line
            ax1.plot([trade.entry_time, trade.exit_time],
                     [trade.entry_spot, trade.exit_spot],
                     color=color, alpha=0.3, linewidth=1, linestyle='--')

    ax1.set_title('NIFTY - EMA Crossover + Option Selling Signals', color='white',
                  fontsize=16, fontweight='bold', pad=15)
    ax1.set_ylabel('Nifty Spot', color='white', fontsize=12)
    ax1.legend(loc='upper left', fontsize=9, facecolor='#1a1a2e', edgecolor='#444',
               labelcolor='white')
    ax1.grid(True, alpha=0.15, color='#666')

    # --- Volume ---
    colors = ['#00e676' if data['close'].iloc[i] >= data['open'].iloc[i] else '#ff1744'
              for i in range(len(data))]
    ax2.bar(data.index, data['volume'], color=colors, alpha=0.6, width=0.005)
    ax2.plot(data.index, data['vol_sma'], color='#ffeb3b', linewidth=1, alpha=0.7, label='Vol SMA')
    ax2.set_ylabel('Volume', color='white', fontsize=12)
    ax2.legend(loc='upper left', fontsize=9, facecolor='#1a1a2e', edgecolor='#444',
               labelcolor='white')
    ax2.grid(True, alpha=0.15, color='#666')

    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight', facecolor='#1a1a2e')
    plt.close()
    print(f"  ✅ Signals chart saved: {save_path}")


def plot_equity_curve(equity_curve, initial_capital, save_path: str = None):
    """Plot equity curve with drawdown."""
    ensure_results_dir()
    if save_path is None:
        save_path = os.path.join(RESULTS_DIR, "equity_curve.png")

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(16, 8), height_ratios=[3, 1],
                                     facecolor='#1a1a2e')

    for ax in [ax1, ax2]:
        ax.set_facecolor('#16213e')
        ax.tick_params(colors='white')
        for spine in ax.spines.values():
            spine.set_color('#444')

    equity = np.array(equity_curve)
    x = range(len(equity))

    # Equity
    ax1.plot(x, equity, color='#00e5ff', linewidth=1.5, label='Equity')
    ax1.axhline(y=initial_capital, color='#666', linestyle='--', alpha=0.5, label='Initial Capital')
    ax1.fill_between(x, initial_capital, equity,
                      where=equity >= initial_capital, color='#00e676', alpha=0.1)
    ax1.fill_between(x, initial_capital, equity,
                      where=equity < initial_capital, color='#ff1744', alpha=0.1)

    ax1.set_title('Equity Curve', color='white', fontsize=14, fontweight='bold')
    ax1.set_ylabel('Capital (₹)', color='white')
    ax1.legend(facecolor='#1a1a2e', edgecolor='#444', labelcolor='white')
    ax1.grid(True, alpha=0.15, color='#666')

    # Drawdown
    peak = np.maximum.accumulate(equity)
    drawdown = (peak - equity) / peak * 100
    ax2.fill_between(x, 0, drawdown, color='#ff1744', alpha=0.4)
    ax2.plot(x, drawdown, color='#ff1744', linewidth=0.8)
    ax2.set_title('Drawdown (%)', color='white', fontsize=12)
    ax2.set_ylabel('DD %', color='white')
    ax2.invert_yaxis()
    ax2.grid(True, alpha=0.15, color='#666')

    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight', facecolor='#1a1a2e')
    plt.close()
    print(f"  ✅ Equity curve saved: {save_path}")


def plot_trade_analysis(trades, save_path: str = None):
    """Plot trade analysis: PnL distribution, win/loss, monthly returns."""
    ensure_results_dir()
    if save_path is None:
        save_path = os.path.join(RESULTS_DIR, "trade_analysis.png")

    if not trades:
        print("  ⚠️ No trades to analyze")
        return

    fig, axes = plt.subplots(2, 2, figsize=(16, 10), facecolor='#1a1a2e')

    for ax in axes.flat:
        ax.set_facecolor('#16213e')
        ax.tick_params(colors='white')
        for spine in ax.spines.values():
            spine.set_color('#444')

    pnls = [t.pnl for t in trades]
    wins = [p for p in pnls if p > 0]
    losses = [p for p in pnls if p <= 0]

    # 1. PnL Distribution
    ax = axes[0, 0]
    ax.hist(pnls, bins=30, color='#00e5ff', alpha=0.7, edgecolor='#1a1a2e')
    ax.axvline(x=0, color='#ff9800', linewidth=2, linestyle='--')
    ax.axvline(x=np.mean(pnls), color='#00e676', linewidth=1.5, linestyle='-', label=f'Mean: ₹{np.mean(pnls):.0f}')
    ax.set_title('PnL Distribution', color='white', fontsize=12, fontweight='bold')
    ax.set_xlabel('PnL (₹)', color='white')
    ax.legend(facecolor='#1a1a2e', edgecolor='#444', labelcolor='white')
    ax.grid(True, alpha=0.15, color='#666')

    # 2. Cumulative PnL
    ax = axes[0, 1]
    cum_pnl = np.cumsum(pnls)
    colors_line = ['#00e676' if p >= 0 else '#ff1744' for p in pnls]
    ax.bar(range(len(pnls)), pnls, color=colors_line, alpha=0.6)
    ax2 = ax.twinx()
    ax2.plot(range(len(cum_pnl)), cum_pnl, color='#ffeb3b', linewidth=2, label='Cumulative')
    ax2.tick_params(colors='white')
    ax.set_title('Trade-wise PnL', color='white', fontsize=12, fontweight='bold')
    ax.set_xlabel('Trade #', color='white')
    ax.set_ylabel('PnL (₹)', color='white')
    ax2.set_ylabel('Cumulative (₹)', color='white')
    ax2.legend(facecolor='#1a1a2e', edgecolor='#444', labelcolor='white', loc='upper left')
    ax.grid(True, alpha=0.15, color='#666')

    # 3. Win/Loss Pie
    ax = axes[1, 0]
    if wins or losses:
        sizes = [len(wins), len(losses)]
        labels = [f'Wins ({len(wins)})', f'Losses ({len(losses)})']
        colors_pie = ['#00e676', '#ff1744']
        wedges, texts, autotexts = ax.pie(sizes, labels=labels, colors=colors_pie,
                                           autopct='%1.1f%%', startangle=90,
                                           textprops={'color': 'white'})
        for t in autotexts:
            t.set_fontsize(12)
            t.set_fontweight('bold')
    ax.set_title('Win/Loss Ratio', color='white', fontsize=12, fontweight='bold')

    # 4. Exit Reason Breakdown
    ax = axes[1, 1]
    exit_reasons = {}
    for t in trades:
        reason = t.status.value
        exit_reasons[reason] = exit_reasons.get(reason, 0) + 1

    if exit_reasons:
        reasons = list(exit_reasons.keys())
        counts = list(exit_reasons.values())
        color_map = {
            'TARGET_HIT': '#00e676',
            'SL_HIT': '#ff1744',
            'TRAILING_SL': '#ff9800',
            'TIME_EXIT': '#2196f3',
            'FORCED_EXIT': '#9c27b0',
        }
        bar_colors = [color_map.get(r, '#666') for r in reasons]
        ax.barh(reasons, counts, color=bar_colors, alpha=0.8)
        ax.set_title('Exit Reasons', color='white', fontsize=12, fontweight='bold')
        ax.set_xlabel('Count', color='white')
    ax.grid(True, alpha=0.15, color='#666')

    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight', facecolor='#1a1a2e')
    plt.close()
    print(f"  ✅ Trade analysis saved: {save_path}")


def print_summary_report(result, initial_capital):
    """Print a formatted summary report to console."""
    summary = result.summary()

    print("\n" + "=" * 60)
    print("  📊 BACKTEST RESULTS SUMMARY")
    print("=" * 60)

    print(f"\n  💰 Capital: ₹{initial_capital:>12,.0f} → ₹{initial_capital + summary['Total PnL (₹)']:>12,.0f}")
    print(f"  📈 Return:  {summary['Total PnL (₹)'] / initial_capital * 100:>11.2f}%")

    # Cost summary
    if result.trades:
        total_gross = sum(t.gross_pnl for t in result.trades)
        total_costs = sum(t.total_costs for t in result.trades)
        print(f"\n  {'─' * 40}")
        print(f"  Gross PnL:       ₹{total_gross:>10,.0f}")
        print(f"  Total Costs:    -₹{total_costs:>10,.0f}")
        print(f"  Net PnL:         ₹{total_gross - total_costs:>10,.0f}")
        print(f"  Avg Cost/Trade:  ₹{total_costs / len(result.trades):>10,.0f}")

    print(f"\n  {'─' * 40}")
    print(f"  Total Trades:     {summary['Total Trades']:>8}")
    print(f"  Winners:          {summary['Winning Trades']:>8}  ({summary['Win Rate (%)']:.1f}%)")
    print(f"  Losers:           {summary['Losing Trades']:>8}")
    print(f"  {'─' * 40}")
    print(f"  Avg Win:         ₹{summary['Avg Win (₹)']:>10,.0f}")
    print(f"  Avg Loss:        ₹{summary['Avg Loss (₹)']:>10,.0f}")
    print(f"  Profit Factor:    {summary['Profit Factor']:>8}")
    print(f"  {'─' * 40}")
    print(f"  Max Drawdown:     {summary['Max Drawdown (%)']:>7}%")
    print(f"  Sharpe Ratio:     {summary['Sharpe Ratio']:>8}")
    print(f"  {'─' * 40}")

    # Trade details
    if result.trades:
        print(f"\n  📋 TRADE LOG (Last 10)")
        print(f"  {'#':>3} {'Entry':>12} {'Dir':>5} {'Opt':>4} {'Strike':>8} "
              f"{'Entry₹':>8} {'Exit₹':>8} {'PnL':>10} {'Status':>12}")
        print(f"  {'─' * 85}")

        for i, t in enumerate(result.trades[-10:], 1):
            dir_str = "BULL" if t.direction.value == 1 else "BEAR"
            pnl_str = f"₹{t.pnl:>+,.0f}"
            entry_str = t.entry_time.strftime('%m/%d %H:%M') if t.entry_time else "N/A"
            print(f"  {i:>3} {entry_str:>12} {dir_str:>5} {t.option_type:>4} "
                  f"{t.strike:>8.0f} {t.entry_premium:>8.1f} {t.exit_premium:>8.1f} "
                  f"{pnl_str:>10} {t.status.value:>12}")

    print("\n" + "=" * 60)


def generate_report(data, result, initial_capital):
    """Generate all charts and print summary."""
    ensure_results_dir()

    print("\n📊 Generating backtest report...")
    print_summary_report(result, initial_capital)

    plot_signals_chart(data, result.trades)
    plot_equity_curve(result.equity_curve, initial_capital)
    plot_trade_analysis(result.trades)

    # Save detailed Excel report
    save_excel_report(data, result, initial_capital)

    print(f"\n  📁 All results saved to: {RESULTS_DIR}/")


def save_excel_report(data: pd.DataFrame, result, initial_capital: float):
    """
    Generate a comprehensive Excel report with multiple sheets:
      1. Summary - Overall performance metrics
      2. Trade Log - Every trade with full details
      3. Daily PnL - Day-wise profit/loss breakdown
      4. Signals - All crossover signals detected
      5. Equity Curve - Bar-by-bar equity data
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    ensure_results_dir()
    xlsx_path = os.path.join(RESULTS_DIR, "backtest_report.xlsx")

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:

        # ── Sheet 1: Summary ──────────────────────────────────────
        summary = result.summary()
        final_capital = initial_capital + summary['Total PnL (₹)']
        summary_data = {
            'Metric': [
                'Strategy', 'Timeframe', 'Data Period',
                '', '── Capital ──',
                'Initial Capital (₹)', 'Final Capital (₹)', 'Net PnL (₹)', 'Return (%)',
                '', '── Trade Stats ──',
                'Total Trades', 'Winning Trades', 'Losing Trades', 'Win Rate (%)',
                '', '── PnL Breakdown ──',
                'Avg Win (₹)', 'Avg Loss (₹)', 'Largest Win (₹)', 'Largest Loss (₹)',
                'Gross Profit (₹)', 'Gross Loss (₹)',
                '', '── Risk Metrics ──',
                'Profit Factor', 'Max Drawdown (%)', 'Sharpe Ratio',
                'Avg Trade Duration',
            ],
            'Value': [
                'EMA 9/21 Crossover + Option Selling', '15 min',
                f"{data.index[0].strftime('%Y-%m-%d')} to {data.index[-1].strftime('%Y-%m-%d')}",
                '', '',
                f'{initial_capital:,.0f}', f'{final_capital:,.0f}',
                f'{summary["Total PnL (₹)"]:,.2f}',
                f'{summary["Total PnL (₹)"] / initial_capital * 100:.2f}%',
                '', '',
                summary['Total Trades'], summary['Winning Trades'],
                summary['Losing Trades'], f'{summary["Win Rate (%)"]:.1f}%',
                '', '',
                f'{summary["Avg Win (₹)"]:,.2f}', f'{summary["Avg Loss (₹)"]:,.2f}',
                f'{max((t.pnl for t in result.trades), default=0):,.2f}',
                f'{min((t.pnl for t in result.trades), default=0):,.2f}',
                f'{sum(t.pnl for t in result.trades if t.pnl > 0):,.2f}',
                f'{sum(t.pnl for t in result.trades if t.pnl <= 0):,.2f}',
                '', '',
                summary['Profit Factor'], f'{summary["Max Drawdown (%)"]}%',
                summary['Sharpe Ratio'],
                _avg_trade_duration(result.trades),
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)

        # ── Sheet 2: Detailed Trade Log ───────────────────────────
        if result.trades:
            trade_rows = []
            cumulative_pnl = 0
            for i, t in enumerate(result.trades, 1):
                cumulative_pnl += t.pnl
                duration = ''
                if t.entry_time and t.exit_time:
                    dur = t.exit_time - t.entry_time
                    hours = dur.total_seconds() / 3600
                    if hours >= 1:
                        duration = f'{hours:.1f} hrs'
                    else:
                        duration = f'{dur.total_seconds() / 60:.0f} min'

                trade_rows.append({
                    'Trade #': i,
                    'Entry Date': t.entry_time.strftime('%Y-%m-%d') if t.entry_time else '',
                    'Entry Time': t.entry_time.strftime('%H:%M:%S') if t.entry_time else '',
                    'Exit Date': t.exit_time.strftime('%Y-%m-%d') if t.exit_time else '',
                    'Exit Time': t.exit_time.strftime('%H:%M:%S') if t.exit_time else '',
                    'Duration': duration,
                    'Direction': 'BULLISH' if t.direction.value == 1 else 'BEARISH',
                    'Signal': 'Sell PUT' if t.direction.value == 1 else 'Sell CALL',
                    'Option Type': t.option_type,
                    'Strike Price': t.strike,
                    'Entry Spot Price': round(t.entry_spot, 2),
                    'Exit Spot Price': round(t.exit_spot, 2),
                    'Spot Move (pts)': round(t.exit_spot - t.entry_spot, 2),
                    'Entry Premium (₹)': round(t.entry_premium, 2),
                    'Exit Premium (₹)': round(t.exit_premium, 2),
                    'Premium Collected (₹)': round(t.entry_premium - t.exit_premium, 2),
                    'Lots': t.lots,
                    'Quantity': t.lots * 25,
                    'Gross PnL (₹)': round(t.gross_pnl, 2),
                    'Brokerage (₹)': round(t.costs.get('brokerage', 0), 2),
                    'Slippage (₹)': round(t.costs.get('slippage', 0), 2),
                    'STT (₹)': round(t.costs.get('stt', 0), 2),
                    'Exchange+GST (₹)': round(t.costs.get('exchange', 0) + t.costs.get('gst', 0), 2),
                    'Total Costs (₹)': round(t.total_costs, 2),
                    'Net PnL (₹)': round(t.pnl, 2),
                    'Cumulative PnL (₹)': round(cumulative_pnl, 2),
                    'Exit Reason': t.status.value.replace('_', ' ').title(),
                    'Result': 'WIN' if t.pnl > 0 else 'LOSS',
                    'Max Favorable (₹)': round(t.max_favorable * t.lots * 25, 2),
                    'SL Spot Level': round(t.sl_spot, 2),
                    'Target Spot Level': round(t.target_spot, 2),
                })

            df_trades = pd.DataFrame(trade_rows)
            df_trades.to_excel(writer, sheet_name='Trade Log', index=False)

        # ── Sheet 3: Daily PnL ────────────────────────────────────
        if result.trades:
            daily_data = {}
            for t in result.trades:
                day = t.entry_time.strftime('%Y-%m-%d') if t.entry_time else 'Unknown'
                if day not in daily_data:
                    daily_data[day] = {'trades': 0, 'wins': 0, 'losses': 0, 'pnl': 0}
                daily_data[day]['trades'] += 1
                daily_data[day]['pnl'] += t.pnl
                if t.pnl > 0:
                    daily_data[day]['wins'] += 1
                else:
                    daily_data[day]['losses'] += 1

            daily_rows = []
            running_pnl = 0
            for day in sorted(daily_data.keys()):
                d = daily_data[day]
                running_pnl += d['pnl']
                daily_rows.append({
                    'Date': day,
                    'Trades': d['trades'],
                    'Wins': d['wins'],
                    'Losses': d['losses'],
                    'Win Rate (%)': round(d['wins'] / d['trades'] * 100, 1) if d['trades'] > 0 else 0,
                    'Day PnL (₹)': round(d['pnl'], 2),
                    'Cumulative PnL (₹)': round(running_pnl, 2),
                    'Day Result': 'GREEN' if d['pnl'] > 0 else 'RED',
                })

            df_daily = pd.DataFrame(daily_rows)
            df_daily.to_excel(writer, sheet_name='Daily PnL', index=False)

        # ── Sheet 4: All Signals Detected ─────────────────────────
        signal_rows = []
        for idx in range(len(data)):
            row = data.iloc[idx]
            is_bull = row.get('crossover_bull', False)
            is_bear = row.get('crossover_bear', False)
            if is_bull or is_bear:
                bar_time = data.index[idx]
                signal_rows.append({
                    'Date': bar_time.strftime('%Y-%m-%d'),
                    'Time': bar_time.strftime('%H:%M'),
                    'Signal Type': 'BULLISH (+)' if is_bull else 'BEARISH (+)',
                    'Close Price': round(row['close'], 2),
                    'EMA Fast (9)': round(row['ema_fast'], 2),
                    'EMA Slow (21)': round(row['ema_slow'], 2),
                    'Close > EMA Slow': 'Yes' if row['close'] > row['ema_slow'] else 'No',
                    'Close < EMA Slow': 'Yes' if row['close'] < row['ema_slow'] else 'No',
                    'High Volume': 'Yes' if row.get('high_volume', False) else 'No',
                    'Trending (ADX>20)': 'Yes' if row.get('not_sideways', False) else 'No',
                    'RSI': round(row.get('rsi', 0), 1),
                    'Volume': int(row['volume']),
                    'Avg Volume': int(row.get('vol_sma', 0)) if not pd.isna(row.get('vol_sma', 0)) else 0,
                    'Tradeable': 'YES' if (
                        (is_bull and row['close'] > row['ema_slow'] and row.get('high_volume', False) and row.get('not_sideways', False)) or
                        (is_bear and row['close'] < row['ema_slow'] and row.get('high_volume', False) and row.get('not_sideways', False))
                    ) else 'NO (filtered)',
                })

        if signal_rows:
            df_signals = pd.DataFrame(signal_rows)
            df_signals.to_excel(writer, sheet_name='All Signals', index=False)

        # ── Sheet 5: Equity Curve Data ────────────────────────────
        equity_rows = []
        step = max(1, len(result.equity_curve) // 500)  # Sample to keep sheet manageable
        for i in range(0, len(result.equity_curve), step):
            if i < len(data):
                bar_time = data.index[i]
                equity_rows.append({
                    'Date': bar_time.strftime('%Y-%m-%d'),
                    'Time': bar_time.strftime('%H:%M'),
                    'Spot Close': round(data.iloc[i]['close'], 2),
                    'Equity (₹)': round(result.equity_curve[i], 2),
                    'Return (%)': round((result.equity_curve[i] - initial_capital) / initial_capital * 100, 3),
                })
        if equity_rows:
            df_equity = pd.DataFrame(equity_rows)
            df_equity.to_excel(writer, sheet_name='Equity Curve', index=False)

        # ── Format all sheets ─────────────────────────────────────
        _format_excel_sheets(writer)

    print(f"  ✅ Excel report saved: {xlsx_path}")


def _avg_trade_duration(trades) -> str:
    """Calculate average trade duration as a readable string."""
    if not trades:
        return 'N/A'
    durations = []
    for t in trades:
        if t.entry_time and t.exit_time:
            dur = (t.exit_time - t.entry_time).total_seconds()
            durations.append(dur)
    if not durations:
        return 'N/A'
    avg_sec = np.mean(durations)
    hours = avg_sec / 3600
    if hours >= 24:
        return f'{hours / 24:.1f} days'
    elif hours >= 1:
        return f'{hours:.1f} hours'
    else:
        return f'{avg_sec / 60:.0f} min'


def _format_excel_sheets(writer):
    """Apply professional formatting to all Excel sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    win_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    loss_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    for sheet_name in writer.book.sheetnames:
        ws = writer.book[sheet_name]

        # Format headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Auto-fit column widths
        for col_idx, col in enumerate(ws.columns, 1):
            max_width = 0
            col_letter = get_column_letter(col_idx)
            for cell in col:
                if cell.value:
                    max_width = max(max_width, len(str(cell.value)))
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[col_letter].width = min(max_width + 3, 25)

        # Color-code WIN/LOSS rows in Trade Log
        if sheet_name == 'Trade Log':
            result_col = None
            for cell in ws[1]:
                if cell.value == 'Result':
                    result_col = cell.column
                    break
            if result_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    result_cell = row[result_col - 1]
                    if result_cell.value == 'WIN':
                        for cell in row:
                            cell.fill = win_fill
                    elif result_cell.value == 'LOSS':
                        for cell in row:
                            cell.fill = loss_fill

        # Color-code Daily PnL
        if sheet_name == 'Daily PnL':
            result_col = None
            for cell in ws[1]:
                if cell.value == 'Day Result':
                    result_col = cell.column
                    break
            if result_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    result_cell = row[result_col - 1]
                    if result_cell.value == 'GREEN':
                        for cell in row:
                            cell.fill = win_fill
                    elif result_cell.value == 'RED':
                        for cell in row:
                            cell.fill = loss_fill

        # Color-code tradeable signals
        if sheet_name == 'All Signals':
            trade_col = None
            for cell in ws[1]:
                if cell.value == 'Tradeable':
                    trade_col = cell.column
                    break
            if trade_col:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    trade_cell = row[trade_col - 1]
                    if trade_cell.value == 'YES':
                        for cell in row:
                            cell.fill = win_fill

        # Freeze top row
        ws.freeze_panes = 'A2'
        # Enable auto-filter
        ws.auto_filter.ref = ws.dimensions
