"""
Visualization & Excel Report for Sapphire Strategy.
Dark-themed charts + professional Excel workbook.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import FancyBboxPatch
import os

# ─── Color Palette ──────────────────────────────────────────────────────────
SAPPHIRE_BG = "#0a0e27"
SAPPHIRE_CARD = "#141938"
SAPPHIRE_BLUE = "#2563eb"
SAPPHIRE_CYAN = "#06b6d4"
SAPPHIRE_GREEN = "#10b981"
SAPPHIRE_RED = "#ef4444"
SAPPHIRE_GOLD = "#f59e0b"
SAPPHIRE_PURPLE = "#8b5cf6"
SAPPHIRE_TEXT = "#e2e8f0"
SAPPHIRE_GRID = "#1e293b"


def setup_dark_style():
    """Apply Sapphire dark theme."""
    plt.rcParams.update({
        "figure.facecolor": SAPPHIRE_BG,
        "axes.facecolor": SAPPHIRE_CARD,
        "axes.edgecolor": SAPPHIRE_GRID,
        "axes.labelcolor": SAPPHIRE_TEXT,
        "text.color": SAPPHIRE_TEXT,
        "xtick.color": SAPPHIRE_TEXT,
        "ytick.color": SAPPHIRE_TEXT,
        "grid.color": SAPPHIRE_GRID,
        "grid.alpha": 0.3,
        "font.family": "sans-serif",
        "font.size": 10,
    })


# ─── Equity Curve ───────────────────────────────────────────────────────────

def plot_equity_curve(result, save_path: str = "results/sapphire_equity.png"):
    """Plot equity curve with drawdown overlay."""
    setup_dark_style()
    
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 8), height_ratios=[3, 1],
                                     gridspec_kw={"hspace": 0.08})
    
    equity = np.array(result.equity_curve)
    x = range(len(equity))
    
    # Equity curve
    ax1.fill_between(x, equity, equity[0], alpha=0.15, color=SAPPHIRE_CYAN)
    ax1.plot(x, equity, color=SAPPHIRE_CYAN, linewidth=1.5, label="Equity")
    
    # Highlight peak
    peak_idx = np.argmax(equity)
    ax1.scatter([peak_idx], [equity[peak_idx]], color=SAPPHIRE_GOLD, s=80, zorder=5,
                label=f"Peak: ₹{equity[peak_idx]:,.0f}")
    
    # Start/end markers
    ax1.scatter([0], [equity[0]], color=SAPPHIRE_PURPLE, s=60, zorder=5)
    ax1.scatter([len(equity)-1], [equity[-1]], color=SAPPHIRE_GREEN if equity[-1] > equity[0] else SAPPHIRE_RED,
                s=60, zorder=5)
    
    ax1.set_title("💎 SAPPHIRE — Equity Curve", fontsize=16, fontweight="bold",
                   color=SAPPHIRE_CYAN, pad=15)
    ax1.set_ylabel("Capital (₹)")
    ax1.legend(loc="upper left", fontsize=9, framealpha=0.3)
    ax1.grid(True, alpha=0.2)
    ax1.set_xticklabels([])
    
    # Drawdown
    peak = np.maximum.accumulate(equity)
    dd = (peak - equity) / peak * 100
    
    ax2.fill_between(x, -dd, 0, alpha=0.4, color=SAPPHIRE_RED)
    ax2.plot(x, -dd, color=SAPPHIRE_RED, linewidth=0.8)
    ax2.set_ylabel("Drawdown (%)")
    ax2.set_xlabel("Trade #")
    ax2.grid(True, alpha=0.2)
    ax2.set_ylim([-max(dd) * 1.3, 0.5])
    
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight", facecolor=SAPPHIRE_BG)
    plt.close()
    print(f"  Equity curve saved: {save_path}")


# ─── Monthly Heatmap ────────────────────────────────────────────────────────

def plot_monthly_heatmap(result, save_path: str = "results/sapphire_monthly.png"):
    """Plot monthly P&L heatmap."""
    setup_dark_style()
    
    if not result.daily_pnl:
        return
    
    df = pd.DataFrame(result.daily_pnl)
    df["date"] = pd.to_datetime(df["date"])
    df["year"] = df["date"].dt.year
    df["month"] = df["date"].dt.month
    
    monthly = df.groupby(["year", "month"])["net_pnl"].sum().reset_index()
    monthly["pnl_pct"] = (monthly["net_pnl"] / result.initial_capital) * 100
    
    years = sorted(monthly["year"].unique())
    months = list(range(1, 13))
    month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    fig, ax = plt.subplots(figsize=(14, max(3, len(years) * 1.2 + 2)))
    
    # Build heatmap data
    data = np.full((len(years), 12), np.nan)
    for _, row in monthly.iterrows():
        y_idx = years.index(int(row["year"]))
        m_idx = int(row["month"]) - 1
        data[y_idx, m_idx] = row["pnl_pct"]
    
    # Color map
    vmax = max(abs(np.nanmin(data)), abs(np.nanmax(data)))
    im = ax.imshow(data, cmap="RdYlGn", aspect="auto", vmin=-vmax, vmax=vmax)
    
    # Labels
    ax.set_xticks(range(12))
    ax.set_xticklabels(month_labels)
    ax.set_yticks(range(len(years)))
    ax.set_yticklabels(years)
    
    # Annotate cells
    for i in range(len(years)):
        for j in range(12):
            val = data[i, j]
            if not np.isnan(val):
                color = "black" if abs(val) < vmax * 0.5 else "white"
                ax.text(j, i, f"{val:+.1f}%", ha="center", va="center",
                       fontsize=10, fontweight="bold", color=color)
    
    ax.set_title("💎 SAPPHIRE — Monthly Returns Heatmap", fontsize=14,
                  fontweight="bold", color=SAPPHIRE_CYAN, pad=15)
    
    plt.colorbar(im, ax=ax, label="Return (%)", shrink=0.8)
    
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight", facecolor=SAPPHIRE_BG)
    plt.close()
    print(f"  Monthly heatmap saved: {save_path}")


# ─── Trade Analysis ─────────────────────────────────────────────────────────

def plot_trade_analysis(result, save_path: str = "results/sapphire_analysis.png"):
    """Plot trade analysis: P&L distribution, win/loss, exit reasons."""
    setup_dark_style()
    
    if not result.trades:
        return
    
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    
    pnls = [t.net_pnl for t in result.trades]
    
    # 1. P&L Distribution
    ax = axes[0, 0]
    colors = [SAPPHIRE_GREEN if p > 0 else SAPPHIRE_RED for p in pnls]
    ax.bar(range(len(pnls)), pnls, color=colors, alpha=0.8, width=1.0)
    ax.axhline(0, color=SAPPHIRE_TEXT, linewidth=0.5, alpha=0.5)
    ax.set_title("Trade-wise P&L", fontsize=12, fontweight="bold")
    ax.set_xlabel("Trade #")
    ax.set_ylabel("P&L (₹)")
    ax.grid(True, alpha=0.2)
    
    # 2. P&L Histogram
    ax = axes[0, 1]
    ax.hist(pnls, bins=30, color=SAPPHIRE_BLUE, alpha=0.7, edgecolor=SAPPHIRE_CARD)
    ax.axvline(np.mean(pnls), color=SAPPHIRE_GOLD, linestyle="--",
               label=f"Mean: ₹{np.mean(pnls):,.0f}")
    ax.axvline(0, color=SAPPHIRE_RED, linewidth=0.8, alpha=0.5)
    ax.set_title("P&L Distribution", fontsize=12, fontweight="bold")
    ax.set_xlabel("P&L (₹)")
    ax.set_ylabel("Frequency")
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.2)
    
    # 3. Exit Reasons Pie
    ax = axes[1, 0]
    reasons = {}
    for t in result.trades:
        reasons[t.exit_reason] = reasons.get(t.exit_reason, 0) + 1
    
    labels = list(reasons.keys())
    sizes = list(reasons.values())
    pie_colors = [SAPPHIRE_CYAN, SAPPHIRE_RED, SAPPHIRE_GOLD, SAPPHIRE_PURPLE,
                  SAPPHIRE_GREEN, SAPPHIRE_BLUE][:len(labels)]
    
    wedges, texts, autotexts = ax.pie(
        sizes, labels=labels, colors=pie_colors, autopct="%1.1f%%",
        textprops={"fontsize": 9, "color": SAPPHIRE_TEXT},
        pctdistance=0.75, startangle=90
    )
    for t in autotexts:
        t.set_fontweight("bold")
    ax.set_title("Exit Reasons", fontsize=12, fontweight="bold")
    
    # 4. Cumulative P&L
    ax = axes[1, 1]
    cum_pnl = np.cumsum(pnls)
    ax.fill_between(range(len(cum_pnl)), cum_pnl, 0, alpha=0.2,
                     color=SAPPHIRE_GREEN if cum_pnl[-1] > 0 else SAPPHIRE_RED)
    ax.plot(cum_pnl, color=SAPPHIRE_CYAN, linewidth=1.5)
    ax.axhline(0, color=SAPPHIRE_TEXT, linewidth=0.5, alpha=0.5)
    ax.set_title("Cumulative P&L", fontsize=12, fontweight="bold")
    ax.set_xlabel("Trade #")
    ax.set_ylabel("Cumulative P&L (₹)")
    ax.grid(True, alpha=0.2)
    
    fig.suptitle("💎 SAPPHIRE — Trade Analysis", fontsize=16, fontweight="bold",
                  color=SAPPHIRE_CYAN, y=1.02)
    
    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight", facecolor=SAPPHIRE_BG)
    plt.close()
    print(f"  Trade analysis saved: {save_path}")


# ─── Excel Report ───────────────────────────────────────────────────────────

def save_excel_report(result, save_path: str = "results/sapphire_report.xlsx"):
    """Save comprehensive Excel report with multiple sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    
    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
        
        # ── Sheet 1: Summary ──
        summary_data = {
            "Metric": [
                "Strategy", "Period", "Initial Capital", "Final Capital",
                "", "--- P&L ---", 
                "Total Trades", "Winners", "Losers", "Win Rate",
                "Gross P&L", "Total Costs", "Net P&L", "ROI",
                "", "--- Risk ---",
                "Max Drawdown (%)", "Max Drawdown (₹)", "Profit Factor", "Sharpe Ratio",
                "", "--- Trades ---",
                "Avg Win", "Avg Loss", "Best Trade", "Worst Trade",
                "Max Consec Wins", "Max Consec Losses",
                "", "--- Monthly ---",
                "Best Month", "Worst Month", "Avg Monthly Return",
            ],
            "Value": [
                "Sapphire Short Strangle",
                f"{len(result.daily_pnl)} trading days",
                f"₹{result.initial_capital:,.0f}",
                f"₹{result.final_capital:,.0f}",
                "", "",
                result.total_trades,
                f"{result.winners} ({result.win_rate:.1f}%)",
                f"{result.losers} ({100-result.win_rate:.1f}%)",
                f"{result.win_rate:.1f}%",
                f"₹{result.total_gross_pnl:,.0f}",
                f"₹{result.total_costs:,.0f}",
                f"₹{result.total_net_pnl:,.0f}",
                f"{result.roi_pct:.1f}%",
                "", "",
                f"{result.max_drawdown_pct:.2f}%",
                f"₹{result.max_drawdown_inr:,.0f}",
                f"{result.profit_factor:.2f}",
                f"{result.sharpe_ratio:.2f}",
                "", "",
                f"₹{result.avg_win:,.0f}",
                f"₹{result.avg_loss:,.0f}",
                f"₹{result.best_trade:,.0f}",
                f"₹{result.worst_trade:,.0f}",
                result.max_consecutive_wins,
                result.max_consecutive_losses,
                "", "",
                f"{result.best_month:.2f}%",
                f"{result.worst_month:.2f}%",
                f"{result.avg_monthly_return:.2f}%",
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)
        
        # ── Sheet 2: Trade Log ──
        if result.trades:
            trade_records = []
            for i, t in enumerate(result.trades, 1):
                trade_records.append({
                    "Trade #": i,
                    "Date": t.date,
                    "Entry Spot": round(t.entry_spot, 2),
                    "Exit Spot": round(t.spot_at_exit, 2),
                    "CE Strike": t.ce_leg.strike if t.ce_leg else "",
                    "PE Strike": t.pe_leg.strike if t.pe_leg else "",
                    "CE Entry Prem": round(t.ce_leg.entry_premium, 2) if t.ce_leg else "",
                    "PE Entry Prem": round(t.pe_leg.entry_premium, 2) if t.pe_leg else "",
                    "CE Exit Prem": round(t.ce_leg.exit_premium, 2) if t.ce_leg else "",
                    "PE Exit Prem": round(t.pe_leg.exit_premium, 2) if t.pe_leg else "",
                    "CE P&L": round(t.ce_leg.net_pnl, 2) if t.ce_leg else "",
                    "PE P&L": round(t.pe_leg.net_pnl, 2) if t.pe_leg else "",
                    "Gross P&L": round(t.gross_pnl, 2),
                    "Costs": round(t.total_costs, 2),
                    "Net P&L": round(t.net_pnl, 2),
                    "Exit Reason": t.exit_reason,
                    "Momentum": "Yes" if t.momentum_detected else "No",
                    "CE Status": t.ce_leg.status.value if t.ce_leg else "",
                    "PE Status": t.pe_leg.status.value if t.pe_leg else "",
                })
            pd.DataFrame(trade_records).to_excel(writer, sheet_name="Trade Log", index=False)
        
        # ── Sheet 3: Daily P&L ──
        if result.daily_pnl:
            df_daily = pd.DataFrame(result.daily_pnl)
            df_daily.to_excel(writer, sheet_name="Daily P&L", index=False)
        
        # ── Sheet 4: Monthly Summary ──
        if result.daily_pnl:
            df = pd.DataFrame(result.daily_pnl)
            df["date"] = pd.to_datetime(df["date"])
            df["month"] = df["date"].dt.to_period("M").astype(str)
            
            monthly = df.groupby("month").agg(
                trades=("net_pnl", "count"),
                gross_pnl=("gross_pnl", "sum"),
                costs=("costs", "sum"),
                net_pnl=("net_pnl", "sum"),
                winners=("net_pnl", lambda x: (x > 0).sum()),
            ).reset_index()
            monthly["win_rate"] = (monthly["winners"] / monthly["trades"] * 100).round(1)
            monthly["return_pct"] = (monthly["net_pnl"] / result.initial_capital * 100).round(2)
            monthly.to_excel(writer, sheet_name="Monthly Summary", index=False)
        
        # ── Sheet 5: Equity Curve ──
        eq_data = pd.DataFrame({
            "Trade #": range(len(result.equity_curve)),
            "Capital": result.equity_curve,
        })
        eq_data.to_excel(writer, sheet_name="Equity Curve", index=False)
        
        # Format all sheets
        _format_excel(writer)
    
    print(f"  Excel report saved: {save_path}")


def _format_excel(writer):
    """Apply professional formatting to Excel sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=10)
    border = Border(
        bottom=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
    )
    
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        
        # Header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        
        # Auto-width columns
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 25)
        
        # Freeze header
        ws.freeze_panes = "A2"
    
    # Highlight P&L cells in Trade Log
    if "Trade Log" in writer.sheets:
        ws = writer.sheets["Trade Log"]
        green_fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
        red_fill = PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid")
        
        # Find Net P&L column
        pnl_col = None
        for cell in ws[1]:
            if cell.value == "Net P&L":
                pnl_col = cell.column
                break
        
        if pnl_col:
            for row in ws.iter_rows(min_row=2, min_col=pnl_col, max_col=pnl_col):
                for cell in row:
                    try:
                        val = float(cell.value) if cell.value else 0
                        cell.fill = green_fill if val > 0 else red_fill
                    except (ValueError, TypeError):
                        pass
