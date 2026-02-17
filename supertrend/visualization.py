"""
Visualization & Excel Report for Supertrend + VWAP Scalping Strategy.
Dark-themed charts + professional Excel workbook.
Teal/Cyan theme to distinguish from EMA (green), Sapphire (blue), Momentum (orange).
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.patches import FancyBboxPatch
import os

# ─── Color Palette (Teal/Cyan Scalping Theme) ───────────────────────────────
ST_BG = "#0a1628"
ST_CARD = "#0f2035"
ST_TEAL = "#14b8a6"
ST_CYAN = "#06b6d4"
ST_GREEN = "#22c55e"
ST_RED = "#ef4444"
ST_GOLD = "#eab308"
ST_PURPLE = "#a855f7"
ST_ORANGE = "#f97316"
ST_TEXT = "#e0f2fe"
ST_GRID = "#1e3a5f"


def setup_dark_style():
    """Apply Supertrend teal/cyan dark theme."""
    plt.rcParams.update({
        "figure.facecolor": ST_BG,
        "axes.facecolor": ST_CARD,
        "axes.edgecolor": ST_GRID,
        "axes.labelcolor": ST_TEXT,
        "text.color": ST_TEXT,
        "xtick.color": ST_TEXT,
        "ytick.color": ST_TEXT,
        "grid.color": ST_GRID,
        "grid.alpha": 0.3,
        "font.family": "sans-serif",
        "font.size": 10,
    })


# ─── Equity Curve ───────────────────────────────────────────────────────────

def plot_equity_curve(result, save_path: str = "results/supertrend_equity.png"):
    """Plot equity curve with drawdown overlay."""
    setup_dark_style()

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(14, 8), height_ratios=[3, 1],
                                     gridspec_kw={"hspace": 0.08})

    equity = np.array(result.equity_curve)
    x = range(len(equity))

    # Equity curve
    ax1.fill_between(x, equity, equity[0], alpha=0.15, color=ST_TEAL)
    ax1.plot(x, equity, color=ST_TEAL, linewidth=1.5, label="Equity")

    # Highlight peak
    peak_idx = np.argmax(equity)
    ax1.scatter([peak_idx], [equity[peak_idx]], color=ST_GOLD, s=80, zorder=5,
                label=f"Peak: \u20b9{equity[peak_idx]:,.0f}")

    # Start/end markers
    ax1.scatter([0], [equity[0]], color=ST_PURPLE, s=60, zorder=5)
    ax1.scatter([len(equity)-1], [equity[-1]],
                color=ST_GREEN if equity[-1] > equity[0] else ST_RED,
                s=60, zorder=5)

    ax1.set_title("\U0001F3AF SUPERTREND + VWAP SCALPING \u2014 Equity Curve", fontsize=16,
                   fontweight="bold", color=ST_TEAL, pad=15)
    ax1.set_ylabel("Capital (\u20b9)")
    ax1.legend(loc="upper left", fontsize=9, framealpha=0.3)
    ax1.grid(True, alpha=0.2)
    ax1.set_xticklabels([])

    # Drawdown
    peak = np.maximum.accumulate(equity)
    dd = (peak - equity) / peak * 100

    ax2.fill_between(x, -dd, 0, alpha=0.4, color=ST_RED)
    ax2.plot(x, -dd, color=ST_RED, linewidth=0.8)
    ax2.set_ylabel("Drawdown (%)")
    ax2.set_xlabel("Trade #")
    ax2.grid(True, alpha=0.2)
    ax2.set_ylim([-max(dd) * 1.3 if max(dd) > 0 else -1, 0.5])

    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight", facecolor=ST_BG)
    plt.close()
    print(f"  Equity curve saved: {save_path}")


# ─── Monthly Heatmap ────────────────────────────────────────────────────────

def plot_monthly_heatmap(result, save_path: str = "results/supertrend_monthly.png"):
    """Plot monthly P&L heatmap."""
    setup_dark_style()

    if not result.daily_pnl:
        return

    df = pd.DataFrame(result.daily_pnl)
    df["date"] = pd.to_datetime(df["date"])
    df["year"] = df["date"].dt.year
    df["month"] = df["date"].dt.month

    monthly = df.groupby(["year", "month"])["net_pnl"].sum().reset_index()
    monthly["pnl_pct"] = (monthly["net_pnl"] / result.initial_capital) * 100

    years = sorted(monthly["year"].unique())
    month_labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    fig, ax = plt.subplots(figsize=(14, max(3, len(years) * 1.2 + 2)))

    data = np.full((len(years), 12), np.nan)
    for _, row in monthly.iterrows():
        y_idx = years.index(int(row["year"]))
        m_idx = int(row["month"]) - 1
        data[y_idx, m_idx] = row["pnl_pct"]

    vmax = max(abs(np.nanmin(data)), abs(np.nanmax(data))) if not np.all(np.isnan(data)) else 1
    im = ax.imshow(data, cmap="RdYlGn", aspect="auto", vmin=-vmax, vmax=vmax)

    ax.set_xticks(range(12))
    ax.set_xticklabels(month_labels)
    ax.set_yticks(range(len(years)))
    ax.set_yticklabels(years)

    for i in range(len(years)):
        for j in range(12):
            val = data[i, j]
            if not np.isnan(val):
                color = "black" if abs(val) < vmax * 0.5 else "white"
                ax.text(j, i, f"{val:+.1f}%", ha="center", va="center",
                       fontsize=10, fontweight="bold", color=color)

    ax.set_title("\U0001F3AF SUPERTREND SCALPING \u2014 Monthly Returns Heatmap", fontsize=14,
                  fontweight="bold", color=ST_TEAL, pad=15)

    plt.colorbar(im, ax=ax, label="Return (%)", shrink=0.8)

    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight", facecolor=ST_BG)
    plt.close()
    print(f"  Monthly heatmap saved: {save_path}")


# ─── Trade Analysis ─────────────────────────────────────────────────────────

def plot_trade_analysis(result, save_path: str = "results/supertrend_analysis.png"):
    """Plot trade analysis: P&L distribution, win/loss, exit reasons, cumulative P&L."""
    setup_dark_style()

    if not result.trades:
        return

    fig, axes = plt.subplots(2, 2, figsize=(14, 10))

    pnls = [t.net_pnl for t in result.trades]

    # 1. P&L Distribution
    ax = axes[0, 0]
    colors = [ST_GREEN if p > 0 else ST_RED for p in pnls]
    ax.bar(range(len(pnls)), pnls, color=colors, alpha=0.8, width=1.0)
    ax.axhline(0, color=ST_TEXT, linewidth=0.5, alpha=0.5)
    ax.set_title("Trade-wise P&L", fontsize=12, fontweight="bold")
    ax.set_xlabel("Trade #")
    ax.set_ylabel("P&L (\u20b9)")
    ax.grid(True, alpha=0.2)

    # 2. P&L Histogram
    ax = axes[0, 1]
    ax.hist(pnls, bins=30, color=ST_TEAL, alpha=0.7, edgecolor=ST_CARD)
    ax.axvline(np.mean(pnls), color=ST_GOLD, linestyle="--",
               label=f"Mean: \u20b9{np.mean(pnls):,.0f}")
    ax.axvline(0, color=ST_RED, linewidth=0.8, alpha=0.5)
    ax.set_title("P&L Distribution", fontsize=12, fontweight="bold")
    ax.set_xlabel("P&L (\u20b9)")
    ax.set_ylabel("Frequency")
    ax.legend(fontsize=9)
    ax.grid(True, alpha=0.2)

    # 3. Exit Reasons Pie
    ax = axes[1, 0]
    reasons = {}
    for t in result.trades:
        reason = t.status.value if hasattr(t.status, "value") else str(t.status)
        reasons[reason] = reasons.get(reason, 0) + 1

    labels = list(reasons.keys())
    sizes = list(reasons.values())
    pie_colors = [ST_GREEN, ST_RED, ST_GOLD, ST_PURPLE,
                  ST_CYAN, ST_ORANGE][:len(labels)]

    wedges, texts, autotexts = ax.pie(
        sizes, labels=labels, colors=pie_colors, autopct="%1.1f%%",
        textprops={"fontsize": 9, "color": ST_TEXT},
        pctdistance=0.75, startangle=90
    )
    for t in autotexts:
        t.set_fontweight("bold")
    ax.set_title("Exit Reasons", fontsize=12, fontweight="bold")

    # 4. Cumulative P&L with target hit markers
    ax = axes[1, 1]
    cum_pnl = np.cumsum(pnls)
    ax.fill_between(range(len(cum_pnl)), cum_pnl, 0, alpha=0.2,
                     color=ST_GREEN if cum_pnl[-1] > 0 else ST_RED)
    ax.plot(cum_pnl, color=ST_TEAL, linewidth=1.5)

    # Highlight target-hit trades
    target_idx = [i for i, t in enumerate(result.trades)
                  if t.status == (t.status.__class__("TARGET_HIT") if hasattr(t.status, 'value') else "TARGET_HIT")]
    if target_idx:
        ax.scatter(target_idx, [cum_pnl[i] for i in target_idx],
                   color=ST_GOLD, s=25, zorder=5, alpha=0.6, label="Target Hit")
        ax.legend(fontsize=9)

    ax.axhline(0, color=ST_TEXT, linewidth=0.5, alpha=0.5)
    ax.set_title("Cumulative P&L", fontsize=12, fontweight="bold")
    ax.set_xlabel("Trade #")
    ax.set_ylabel("Cumulative P&L (\u20b9)")
    ax.grid(True, alpha=0.2)

    fig.suptitle("\U0001F3AF SUPERTREND + VWAP SCALPING \u2014 Trade Analysis", fontsize=16,
                  fontweight="bold", color=ST_TEAL, y=1.02)

    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight", facecolor=ST_BG)
    plt.close()
    print(f"  Trade analysis saved: {save_path}")


# ─── Excel Report ───────────────────────────────────────────────────────────

def save_excel_report(result, save_path: str = "results/supertrend_report.xlsx"):
    """Save comprehensive Excel report with multiple sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(save_path), exist_ok=True)

    with pd.ExcelWriter(save_path, engine="openpyxl") as writer:

        # ── Sheet 1: Summary ──
        summary_data = {
            "Metric": [
                "Strategy", "Period", "Initial Capital", "Final Capital",
                "", "--- P&L ---",
                "Total Trades", "Winners", "Losers", "Win Rate",
                "Gross P&L", "Total Costs", "Net P&L", "ROI",
                "", "--- Risk ---",
                "Max Drawdown (%)", "Max Drawdown (\u20b9)", "Profit Factor", "Sharpe Ratio",
                "", "--- Trades ---",
                "Avg Win", "Avg Loss", "Best Trade", "Worst Trade",
                "Max Consec Wins", "Max Consec Losses",
                "", "--- Scalping ---",
                "Avg Trades/Day", "Target Hit Rate",
                "", "--- Monthly ---",
                "Best Month", "Worst Month", "Avg Monthly Return",
            ],
            "Value": [
                "Supertrend + VWAP Scalping",
                f"{len(set(d['date'] for d in result.daily_pnl))} trading days" if result.daily_pnl else "0 days",
                f"\u20b9{result.initial_capital:,.0f}",
                f"\u20b9{result.final_capital:,.0f}",
                "", "",
                result.total_trades,
                f"{result.winners} ({result.win_rate:.1f}%)",
                f"{result.losers} ({100-result.win_rate:.1f}%)" if result.win_rate else "0",
                f"{result.win_rate:.1f}%",
                f"\u20b9{result.total_gross_pnl:,.0f}",
                f"\u20b9{result.total_costs:,.0f}",
                f"\u20b9{result.total_net_pnl:,.0f}",
                f"{result.roi_pct:.1f}%",
                "", "",
                f"{result.max_drawdown_pct:.2f}%",
                f"\u20b9{result.max_drawdown_inr:,.0f}",
                f"{result.profit_factor:.2f}",
                f"{result.sharpe_ratio:.2f}",
                "", "",
                f"\u20b9{result.avg_win:,.0f}",
                f"\u20b9{result.avg_loss:,.0f}",
                f"\u20b9{result.best_trade:,.0f}",
                f"\u20b9{result.worst_trade:,.0f}",
                result.max_consecutive_wins,
                result.max_consecutive_losses,
                "", "",
                f"{result.avg_trades_per_day:.1f}",
                f"{result.target_hit_rate:.1f}%",
                "", "",
                f"{result.best_month:.2f}%",
                f"{result.worst_month:.2f}%",
                f"{result.avg_monthly_return:.2f}%",
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

        # ── Sheet 2: Trade Log ──
        if result.trades:
            trade_records = []
            for i, t in enumerate(result.trades, 1):
                direction = t.direction.name if hasattr(t.direction, "name") else str(t.direction)
                status = t.status.value if hasattr(t.status, "value") else str(t.status)
                trade_records.append({
                    "Trade #": i,
                    "Date": str(t.entry_time.date()) if t.entry_time else "",
                    "Direction": direction,
                    "Entry Price": round(t.entry_price, 2),
                    "Exit Price": round(t.exit_price, 2),
                    "Lots": t.lots,
                    "SL": round(t.sl_price, 2),
                    "Target": round(t.target_price, 2),
                    "Trailing SL": round(t.trailing_sl, 2),
                    "Max Favorable": round(t.max_favorable, 2),
                    "Gross P&L": round(t.gross_pnl, 2),
                    "Costs": round(t.costs, 2),
                    "Net P&L": round(t.net_pnl, 2),
                    "Status": status,
                    "Supertrend": round(t.entry_supertrend, 2),
                    "VWAP": round(t.entry_vwap, 2),
                    "EMA-9": round(t.entry_ema, 2),
                })
            pd.DataFrame(trade_records).to_excel(writer, sheet_name="Trade Log", index=False)

        # ── Sheet 3: Daily P&L ──
        if result.daily_pnl:
            df_daily = pd.DataFrame(result.daily_pnl)
            df_daily.to_excel(writer, sheet_name="Daily P&L", index=False)

        # ── Sheet 4: Monthly Summary ──
        if result.daily_pnl:
            df = pd.DataFrame(result.daily_pnl)
            df["date"] = pd.to_datetime(df["date"])
            df["month"] = df["date"].dt.to_period("M").astype(str)

            monthly = df.groupby("month").agg(
                trades=("net_pnl", "count"),
                gross_pnl=("gross_pnl", "sum"),
                costs=("costs", "sum"),
                net_pnl=("net_pnl", "sum"),
                winners=("net_pnl", lambda x: (x > 0).sum()),
            ).reset_index()
            monthly["win_rate"] = (monthly["winners"] / monthly["trades"] * 100).round(1)
            monthly["return_pct"] = (monthly["net_pnl"] / result.initial_capital * 100).round(2)
            monthly.to_excel(writer, sheet_name="Monthly Summary", index=False)

        # ── Sheet 5: Equity Curve ──
        eq_data = pd.DataFrame({
            "Trade #": range(len(result.equity_curve)),
            "Capital": result.equity_curve,
        })
        eq_data.to_excel(writer, sheet_name="Equity Curve", index=False)

        # Format all sheets
        _format_excel(writer)

    print(f"  Excel report saved: {save_path}")


def _format_excel(writer):
    """Apply professional formatting to Excel sheets."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="0f766e", end_color="0f766e", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Calibri", size=10)
    border = Border(
        bottom=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
    )

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]

        # Header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data formatting
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        # Auto-width columns
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 25)

        # Freeze header
        ws.freeze_panes = "A2"

    # Highlight P&L cells in Trade Log
    if "Trade Log" in writer.sheets:
        ws = writer.sheets["Trade Log"]
        green_fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
        red_fill = PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid")

        pnl_col = None
        for cell in ws[1]:
            if cell.value == "Net P&L":
                pnl_col = cell.column
                break

        if pnl_col:
            for row in ws.iter_rows(min_row=2, min_col=pnl_col, max_col=pnl_col):
                for cell in row:
                    try:
                        val = float(cell.value) if cell.value else 0
                        cell.fill = green_fill if val > 0 else red_fill
                    except (ValueError, TypeError):
                        pass
